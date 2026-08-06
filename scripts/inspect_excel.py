import os
import sys

try:
    import openpyxl
except Exception as e:
    print('IMPORT_ERROR', e)
    sys.exit(1)

path = r'd:\AI_TEST\SunnyApps\etc\챙이네공방.xlsx'
print('exists', os.path.exists(path))
wb = openpyxl.load_workbook(path, data_only=True)
print('sheets', wb.sheetnames)
for ws in wb.worksheets:
    print('---', ws.title, 'rows', ws.max_row, 'cols', ws.max_column)
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 5), values_only=True):
        print(row)
