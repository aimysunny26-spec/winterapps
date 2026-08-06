import json
import mimetypes
import os
from pathlib import Path
from urllib.parse import urlparse, parse_qs

from http.server import ThreadingHTTPServer, BaseHTTPRequestHandler

try:
    import openpyxl
except ImportError as exc:
    raise SystemExit(f"openpyxl is required. Install it with: pip install openpyxl\n{exc}")

ROOT = Path(__file__).resolve().parent
WEB_ROOT = ROOT / "webapp"
WORKBOOK_PATH = ROOT / "etc" / "챙이네공방.xlsx"


def normalize_header(value):
    if value is None:
        return ""
    return str(value).strip().lower().replace("-", "").replace("_", "").replace(" ", "")


def row_to_item(headers, row):
    item = {}
    for idx, header in enumerate(headers):
        item[normalize_header(header)] = row[idx]

    # normalize aliases for the page UI
    title = item.get("title") or item.get("name") or item.get("subject")
    category = item.get("category") or item.get("type") or item.get("tag")
    desc = item.get("desc") or item.get("description") or item.get("content") or item.get("detail")
    content = item.get("content") or item.get("desc") or item.get("description")
    price = item.get("price") or item.get("cost")
    image_url = item.get("imageurl") or item.get("image") or item.get("image_url") or item.get("img")
    author = item.get("author") or item.get("writer") or item.get("user")
    date = item.get("date") or item.get("created") or item.get("postdate")
    active = item.get("active")

    result = {}
    if title is not None:
        result["title"] = title
    if category is not None:
        result["category"] = category
    if desc is not None:
        result["desc"] = desc
    if content is not None:
        result["content"] = content
    if price is not None:
        result["price"] = price
    if image_url is not None:
        result["imageUrl"] = image_url
    if author is not None:
        result["author"] = author
    if date is not None:
        result["date"] = date
    if active is not None:
        result["active"] = active

    # preserve any extra columns too
    for key, value in item.items():
        if key not in result:
            result[key] = value

    return result


def load_sheet(sheet_name):
    if not WORKBOOK_PATH.exists():
        return []

    workbook = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True, read_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            return []

        sheet = workbook[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []

        headers = [str(cell or "") for cell in rows[0]]
        data = []
        for row in rows[1:]:
            if not any(cell is not None and str(cell).strip() != "" for cell in row):
                continue
            item = row_to_item(headers, row)
            if item.get("active") is False or item.get("active") == "FALSE" or item.get("active") == "false":
                continue
            data.append(item)
        return data
    finally:
        workbook.close()


class Handler(BaseHTTPRequestHandler):
    def do_GET(self):
        parsed = urlparse(self.path)
        path = parsed.path

        if path.startswith("/api/"):
            self.handle_api(parsed)
            return

        if path == "/":
            target = WEB_ROOT / "index.html"
        else:
            target = WEB_ROOT / path.lstrip("/")

        if target.is_dir():
            target = target / "index.html"

        if target.exists() and target.is_file():
            content = target.read_bytes()
            content_type = mimetypes.guess_type(str(target))[0] or "application/octet-stream"
            self.send_response(200)
            self.send_header("Content-Type", content_type)
            self.send_header("Content-Length", str(len(content)))
            self.end_headers()
            self.wfile.write(content)
        else:
            self.send_response(404)
            self.end_headers()
            self.wfile.write(b"Not Found")

    def handle_api(self, parsed):
        query = parse_qs(parsed.query)
        if parsed.path == "/api/sheet":
            sheet_name = query.get("name", [""])[0]
            if not sheet_name:
                self.send_json(400, {"error": "Missing name query parameter"})
                return
            self.send_json(200, load_sheet(sheet_name))
            return

        if parsed.path == "/api/sheets":
            workbook = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True, read_only=True)
            try:
                self.send_json(200, workbook.sheetnames)
            finally:
                workbook.close()
            return

        self.send_json(404, {"error": "Not found"})

    def send_json(self, status, payload):
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def log_message(self, format, *args):
        return


if __name__ == "__main__":
    port = int(os.environ.get("PORT", "8000"))
    server = ThreadingHTTPServer(("0.0.0.0", port), Handler)
    print(f"Serving {WEB_ROOT} at http://127.0.0.1:{port}")
    server.serve_forever()
